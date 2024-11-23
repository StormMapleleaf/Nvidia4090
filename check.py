import os
import re
import openpyxl

def check_txt_files(directory, excel_file):
    """
    检查目录中所有的txt文件，验证每一行是否符合{"input":"文本内容","output":"文本内容"}格式。
    冒号后允许有空格或无空格。如果没有问题，输出文件名及无问题信息。
    同时提取班级和姓名并更新Excel文件。
    
    :param directory: 要检查的目录路径
    :param excel_file: Excel 文件路径，存储班级和姓名
    """
    # 加载 Excel 文件
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    # 获取现有数据（避免重复写入）
    existing_data = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        existing_data.add((row[0], row[1]))  # 班级和姓名在 Excel 中的列位置

    # 定义符合要求的正则表达式（支持冒号后有无空格）
    pattern = re.compile(r'^\{"input"\s*:\s*"[^"]*"\s*,\s*"output"\s*:\s*"[^"]*"\}$')
    
    # 遍历目录中的所有txt文件
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.txt'):
                file_path = os.path.join(root, file)
                file_is_valid = True  # 默认假设文件格式无问题
                class_name, student_name = extract_class_and_name(file)  # 提取班级和姓名
                
                if file_is_valid:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        for line_number, line in enumerate(f, start=1):
                            line = line.strip()
                            # 检查是否符合格式
                            if not pattern.match(line):
                                print(f"文件 '{file_path}' 第 {line_number} 行格式有误：{line}")
                                file_is_valid = False
                                break  # 找到第一个问题行后就不再继续检查此文件
                    
                    # 如果文件格式无问题，写入到 Excel 文件
                    if file_is_valid:
                        print(f"文件 '{file}' 无问题")
                        if (class_name, student_name) not in existing_data:
                            sheet.append([class_name, student_name, 1])  # 插入班级、姓名和语料作业列
                            existing_data.add((class_name, student_name))  # 防止重复写入
    
    # 保存 Excel 文件
    wb.save(excel_file)
    print(f"更新完成，保存到 {excel_file}")

def extract_class_and_name(filename):
    """
    从文件名中提取班级和姓名。假设文件名格式为 '班级-姓名.txt'。
    :param filename: 文件名
    :return: (班级, 姓名)
    """
    base_name = os.path.splitext(filename)[0]  # 获取不带扩展名的文件名
    class_name, student_name = base_name.split('-')  # 假设班级和姓名通过'-'分隔
    return class_name, student_name

# 使用方法
directory_to_check = "./workcheck"  # 替换为你的目录路径
excel_file_path = "名单.xlsx"  # 替换为你要保存的 Excel 文件路径
check_txt_files(directory_to_check, excel_file_path)
